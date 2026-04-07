"""
Utility functions for data processing, reporting, and file handling
"""
import secrets
import string
from io import BytesIO
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.database import SessionLocal
from app.models import FleetRecord, SystemSetting
from app.schemas import (
    AnalyticsResponse,
    Anomaly,
    DailySubtotal,
    DashboardStats,
    FleetRecordOut,
    FleetSummary,
)


def generate_account_id(role: str = "user") -> str:
    """
    Generate a random account ID with letters, numbers, and special chars.
    Prefix depends on role: ADM for admin, USE for user.
    """
    alphabet = string.ascii_uppercase + string.digits + "#@&"
    prefix = "ADM-" if role == "admin" else "USE-"
    return prefix + "".join(secrets.choice(alphabet) for _ in range(8))


def get_system_config() -> dict:
    """Fetch system settings as a dict."""
    db = SessionLocal()
    try:
        settings = db.query(SystemSetting).all()
        return {s.key: s.value for s in settings}
    finally:
        db.close()


def calculate_remittance(revenue: float, fleet_code: str, config: dict = None) -> float:
    """
    Calculate remittance based on fleet code prefix.
    Supports dynamic config or defaults to hardcoded rules.
    """
    fleet_code = str(fleet_code).strip()

    if config:
        prefix = fleet_code[0] if fleet_code else ""
        rate_key = f"REMITTANCE_{prefix}"
        if rate_key in config:
            try:
                rate = float(config[rate_key]) / 100
                return revenue * rate
            except (ValueError, TypeError):
                pass

    if fleet_code.startswith("1"):
        return revenue * 0.84
    elif fleet_code.startswith("2"):
        return revenue * 0.875
    return revenue


class DataProcessor:
    """Handles data transformation and aggregation"""

    @staticmethod
    def detect_anomalies(df: pd.DataFrame) -> List[Anomaly]:
        """Detect unusual patterns in fleet data using Z-score analysis."""
        anomalies = []

        if df.empty:
            return anomalies

        # Create a copy to avoid SettingWithCopyWarning
        df = df.copy()
        
        # Calculate group means and std devs using vectorization
        groups = df.groupby("fleet")
        df["group_count"] = groups["amount"].transform("count")
        df["group_std"] = groups["amount"].transform("std")
        
        # Filter valid records directly
        valid_mask = (df["group_count"] >= 5) & (df["group_std"] > 0)
        valid_df = df[valid_mask].copy()
        
        if valid_df.empty:
            return anomalies
            
        valid_df["group_mean"] = valid_df.groupby("fleet")["amount"].transform("mean")
        valid_df["z_score"] = abs(valid_df["amount"] - valid_df["group_mean"]) / valid_df["group_std"]
        
        # Extract high anomalies
        high_anomalies = valid_df[valid_df["z_score"] > 3]
        for _, row in high_anomalies.iterrows():
            anomalies.append(Anomaly(
                date=row["date"],
                fleet=row["fleet"],
                amount=row["amount"],
                reason=f"Significant deviation (Z-score: {row['z_score']:.2f})",
                severity="high",
            ))
            
        # Extract medium anomalies
        medium_anomalies = valid_df[(valid_df["z_score"] > 2) & (valid_df["z_score"] <= 3)]
        for _, row in medium_anomalies.iterrows():
            anomalies.append(Anomaly(
                date=row["date"],
                fleet=row["fleet"],
                amount=row["amount"],
                reason="Unusual amount for this fleet",
                severity="medium",
            ))

        return anomalies

    @staticmethod
    def process_analytics(records: List[FleetRecord]) -> AnalyticsResponse:
        """Process raw records into full analytics response."""
        if not records:
            return AnalyticsResponse(
                records=[],
                fleet_summaries=[],
                daily_subtotals=[],
                dashboard_stats=DashboardStats(
                    total_revenue=0,
                    total_records=0,
                    top_performing_fleet="N/A",
                    average_trip_revenue=0,
                ),
                anomalies=[],
            )

        # Use list comprehension with tuples for vastly superior DataFrame creation speed
        data = [
            (
                r.id, 
                r.date, 
                "2010" if r.fleet.strip().upper() == "2010M" else r.fleet.strip().upper(), 
                r.amount
            ) 
            for r in records
        ]

        df = pd.DataFrame(data, columns=["id", "date", "fleet", "amount"])
        # date is already a python date object coming from SQLAlchemy (model uses Date)
        # However, to be safe, avoid pd.to_datetime if not needed, but we keep it and dt.date to be sure
        df["date"] = pd.to_datetime(df["date"]).dt.date

        config = get_system_config()

        # Fleet Summaries
        fleet_grp = df.groupby("fleet")["amount"].agg(["sum", "count"]).reset_index()
        summaries = [
            FleetSummary(
                fleet=str(row["fleet"]).strip(),
                total_amount=row["sum"],
                record_count=row["count"],
                remittance=calculate_remittance(row["sum"], str(row["fleet"]).strip(), config),
            )
            for _, row in fleet_grp.iterrows()
        ]

        # Daily Subtotals
        daily_grp = (
            df.groupby(["date", "fleet"])["amount"]
            .agg(["sum", "count"])
            .reset_index()
            .sort_values(by=["date", "fleet"])
        )
        subtotals = [
            DailySubtotal(date=row["date"], fleet=row["fleet"], daily_total=row["sum"], pax=row["count"])
            for _, row in daily_grp.iterrows()
        ]

        # Dashboard Stats with Trend
        total_rev = df["amount"].sum()
        total_count = len(df)
        avg_rev = total_rev / total_count if total_count > 0 else 0
        top_fleet = fleet_grp.loc[fleet_grp["sum"].idxmax(), "fleet"] if not fleet_grp.empty else "N/A"

        max_date = df["date"].max()
        rev_last = df[df["date"] >= (max_date - pd.Timedelta(days=7))]["amount"].sum()
        rev_prev = df[
            (df["date"] < (max_date - pd.Timedelta(days=7))) &
            (df["date"] >= (max_date - pd.Timedelta(days=14)))
        ]["amount"].sum()
        trend_percent = ((rev_last - rev_prev) / rev_prev * 100) if rev_prev > 0 else 0.0

        stats = DashboardStats(
            total_revenue=total_rev,
            total_records=total_count,
            top_performing_fleet=top_fleet,
            average_trip_revenue=avg_rev,
            revenue_trend_percent=trend_percent,
        )

        anomalies = DataProcessor.detect_anomalies(df)
        # Limit records sent in JSON to 2000 to prevent massive payloads crashing the browser
        record_objs = [
            FleetRecordOut(id=t[0], date=t[1], fleet=t[2], amount=t[3])
            for t in data[:2000]
        ]

        return AnalyticsResponse(
            records=record_objs,
            fleet_summaries=summaries,
            daily_subtotals=subtotals,
            dashboard_stats=stats,
            anomalies=anomalies,
        )


class ReportGenerator:
    """Generates Excel and PDF reports"""

    @staticmethod
    def generate_excel(analytics: AnalyticsResponse) -> BytesIO:
        """Create a multi-sheet Excel report with custom styling."""
        if not analytics.records:
            output = BytesIO()
            pd.DataFrame().to_excel(output)
            output.seek(0)
            return output

        records_data = [r.model_dump() for r in analytics.records]
        df = pd.DataFrame(records_data)
        df["date"] = pd.to_datetime(df["date"]).dt.date
        df["amount"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0)
        df["fleet"] = df["fleet"].astype(str)

        # Daily Subtotals
        grouped = (
            df.groupby(["date", "fleet"])
            .agg(TotalAmount=("amount", "sum"), FleetCount=("fleet", "count"))
            .reset_index()
            .sort_values(by=["date", "fleet"])
        )

        formatted_rows = []
        for date_val, group in grouped.groupby("date"):
            for _, row in group.iterrows():
                formatted_rows.append({
                    "Date": row["date"],
                    "BUS CODE": row["fleet"],
                    "PAX": row["FleetCount"],
                    "REVENUE": row["TotalAmount"],
                })
            formatted_rows.append({
                "Date": date_val,
                "BUS CODE": "Subtotal",
                "PAX": group["FleetCount"].sum(),
                "REVENUE": group["TotalAmount"].sum(),
            })

        subtotal_df = pd.DataFrame(formatted_rows)

        # Bus Code Performance
        bus_summary_df = (
            df.groupby("fleet")
            .agg(PAX=("fleet", "count"), REVENUE=("amount", "sum"))
            .reset_index()
            .rename(columns={"fleet": "BUS CODE"})
            .sort_values("BUS CODE")
        )

        config = get_system_config()

        def calculate_metrics(row):
            code = str(row["BUS CODE"]).strip()
            remittance = calculate_remittance(row["REVENUE"], code, config)
            fuel_used = row["REVENUE"] * 0.30
            return pd.Series([remittance, fuel_used])

        bus_summary_df[["REMITTANCE", "FUEL USED"]] = bus_summary_df.apply(calculate_metrics, axis=1)
        bus_summary_df.loc[len(bus_summary_df)] = [
            "Grand Total",
            bus_summary_df["PAX"].sum(),
            bus_summary_df["REVENUE"].sum(),
            bus_summary_df["REMITTANCE"].sum(),
            bus_summary_df["FUEL USED"].sum(),
        ]

        output = BytesIO()
        writer = pd.ExcelWriter(output, engine="openpyxl")
        bus_summary_df.to_excel(writer, sheet_name="Bus Performance", index=False)
        subtotal_df.to_excel(writer, sheet_name="Daily Subtotals", index=False)
        writer.close()

        # Apply Styling
        output.seek(0)
        wb = load_workbook(output)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        subtotal_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        def style_worksheet(ws, sheet_name):
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            is_bus_perf = sheet_name == "Bus Performance"
            is_daily = sheet_name == "Daily Subtotals"

            # Cache header names once for number formatting lookups
            header_names = {cell.column: str(cell.value).upper() for cell in ws[1]}

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border

                bus_code = str(row[0].value) if is_bus_perf else (str(row[1].value) if is_daily else None)

                if bus_code in ("Grand Total", "Subtotal"):
                    for cell in row:
                        cell.font = bold_font
                        cell.fill = subtotal_fill
                elif is_bus_perf and bus_code:
                    if bus_code.startswith("1"):
                        for cell in row:
                            cell.fill = yellow_fill
                    elif bus_code.startswith("2"):
                        for cell in row:
                            cell.fill = blue_fill

                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        header_str = header_names.get(cell.column, "")
                        if any(kw in header_str for kw in ("REVENUE", "REMITTANCE", "AMOUNT", "FUEL")):
                            cell.number_format = "#,##0.00"
                        elif any(kw in header_str for kw in ("PAX", "COUNT")):
                            cell.number_format = "#,##0"

            for column_cells in ws.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                    default=0,
                )
                ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

        for sheet_name in wb.sheetnames:
            style_worksheet(wb[sheet_name], sheet_name)

        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)
        return final_output

    @staticmethod
    def generate_pdf(analytics: AnalyticsResponse) -> BytesIO:
        """Create a PDF report."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        elements.append(Paragraph("Fleet Reporting System - Analytical Report", styles["Title"]))
        elements.append(Spacer(1, 12))

        elements.append(Paragraph("Executive Summary", styles["Heading2"]))
        stats = analytics.dashboard_stats
        summary_data = [
            ["Metric", "Value"],
            ["Total Revenue", f"{stats.total_revenue:,.2f}"],
            ["Total Records", f"{stats.total_records}"],
            ["Top Fleet", stats.top_performing_fleet],
            ["Avg Revenue/Trip", f"{stats.average_trip_revenue:,.2f}"],
        ]

        t = Table(summary_data)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("Fleet Performance Breakdown", styles["Heading2"]))
        if analytics.fleet_summaries:
            data = [["Fleet", "Total Revenue", "Count"]]
            for s in analytics.fleet_summaries:
                data.append([s.fleet, f"{s.total_amount:,.2f}", str(s.record_count)])

            t2 = Table(data)
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.blue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ]))
            elements.append(t2)
        else:
            elements.append(Paragraph("No data available", styles["Normal"]))

        doc.build(elements)
        output.seek(0)
        return output
